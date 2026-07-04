#!/usr/bin/env python
"""Exports the full `public.jobs` table (every column) to an .xlsx workbook
for manual review - the Python equivalent of the old n8n
"06-daily-sheet-export" workflow, but covering every field instead of a
curated subset, and run on demand rather than daily.

Usage:
    python export_sheet.py [output_path.xlsx]
"""

import sys
from datetime import datetime, timezone
from pathlib import Path

import httpx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from jobscraper import config

PAGE_SIZE = 1000
TRUNCATE_LONG_TEXT_AT = 400  # keep the sheet readable; full data stays in Supabase


def fetch_all_jobs() -> list[dict]:
    config.require_supabase()
    url = f"{config.SUPABASE_URL}/rest/v1/jobs?select=*&order=created_at.desc"
    headers = {
        "apikey": config.SUPABASE_SERVICE_ROLE_KEY,
        "Authorization": f"Bearer {config.SUPABASE_SERVICE_ROLE_KEY}",
    }
    jobs: list[dict] = []
    offset = 0
    with httpx.Client(timeout=60.0) as client:
        while True:
            page_headers = {
                **headers,
                "Range-Unit": "items",
                "Range": f"{offset}-{offset + PAGE_SIZE - 1}",
            }
            resp = client.get(url, headers=page_headers)
            resp.raise_for_status()
            page = resp.json()
            jobs.extend(page)
            if len(page) < PAGE_SIZE:
                break
            offset += PAGE_SIZE
    return jobs


def _cell_value(value):
    if isinstance(value, list):
        return ", ".join(str(v) for v in value)
    if isinstance(value, str) and len(value) > TRUNCATE_LONG_TEXT_AT:
        return value[:TRUNCATE_LONG_TEXT_AT] + "..."
    return value


def build_workbook(jobs: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "jobs"

    if not jobs:
        ws.append(["No jobs found"])
        return wb

    columns = list(jobs[0].keys())
    ws.append(columns)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    suspicious_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    inactive_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for job in jobs:
        row = [_cell_value(job.get(col)) for col in columns]
        ws.append(row)
        row_idx = ws.max_row
        if job.get("flagged_suspicious"):
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = suspicious_fill
        elif not job.get("is_active", True):
            for col_idx in range(1, len(columns) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = inactive_fill

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(
            [len(col_name)] + [len(str(_cell_value(j.get(col_name)) or "")) for j in jobs[:200]]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    summary = wb.create_sheet("summary")
    total = len(jobs)
    active = sum(1 for j in jobs if j.get("is_active"))
    suspicious = sum(1 for j in jobs if j.get("flagged_suspicious"))
    entry_level = sum(1 for j in jobs if j.get("entry_level"))
    sources: dict[str, int] = {}
    for j in jobs:
        src = (j.get("source") or "unknown").split(":")[0]
        sources[src] = sources.get(src, 0) + 1

    summary.append(["Generated at (UTC)", datetime.now(timezone.utc).isoformat()])
    summary.append(["Total jobs", total])
    summary.append(["Active jobs", active])
    summary.append(["Discontinued jobs", total - active])
    summary.append(["Flagged suspicious", suspicious])
    summary.append(["Entry-level", entry_level])
    summary.append([])
    summary.append(["Source", "Count"])
    for src, count in sorted(sources.items(), key=lambda kv: -kv[1]):
        summary.append([src, count])
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 40

    wb.move_sheet("summary", offset=-len(wb.sheetnames))
    return wb


def main() -> int:
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(__file__).parent / "jobs_export.xlsx"

    print("Fetching all jobs from Supabase...")
    jobs = fetch_all_jobs()
    print(f"Fetched {len(jobs)} rows across {len(jobs[0].keys()) if jobs else 0} columns")

    wb = build_workbook(jobs)
    wb.save(out_path)
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
